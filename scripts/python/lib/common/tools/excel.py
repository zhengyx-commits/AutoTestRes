import logging
import os
import shutil
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from lib.common.system.ADB import ADB
from tools.DB import DB
from tools.yamlTool import yamlTool

adb_tool = ADB()
config_compatibility_yaml = yamlTool(os.getcwd() + '/config/config_ott_hybrid_compatibility.yaml')
resource_url = config_compatibility_yaml.get_note("conf_resource_server").get("resource_url")

class ExcelCreator:
    def __init__(self, data, headers="", version_info="", pipline_type=""):
        self.data = data
        self.left_data = []
        self.headers = headers
        self.excel_path = '/home/amlogic'
        self.excel_name = f'compatibility.xlsx'
        self.excel_file = os.path.join(self.excel_path, self.excel_name)
        self.workbook = None
        self.sheet = None
        self.version_info = version_info
        self.pipline_type = pipline_type

    # create excel
    def create_excel(self):
        existing_files = [file for file in os.listdir(self.excel_path) if file.endswith('.xlsx')]
        if existing_files:
            logging.info(existing_files)
            logging.info(f'Excel file "{self.excel_name}" already exists.')
            self.workbook = load_workbook(self.excel_file)
            self.sheet = self.workbook.active

            # Check if the last video has been fully detected
            last_row = self.sheet.max_row
            status_value = self.sheet.cell(row=last_row, column=11).value
            # with DB() as db:
            #     field_names = ['STATUS']
            #     status_value = db.get_data_fields('COMPATIBILITY', field_names, condition=f'VERSION={self.version_info} AND PIPLINE_TYPE={self.pipline_type}')
            if status_value == 0 or status_value == 1 or status_value == "None":
                # Traverse each row to check if the status column is None
                max_row = self.sheet.max_row
                current_row = self.find_current_test_row()
                for row in range(current_row, max_row + 1):
                    video_value = self.sheet.cell(row=row, column=2).value
                    status_value = self.sheet.cell(row=row, column=11).value
                    if status_value != 0 and status_value != 1 and status_value != "None":
                        self.left_data.append(video_value)
                if len(self.left_data) > 0:
                    logging.info("Continue testing the missed videos")
                    return
                else:
                    logging.info("adding new video rows")
            else:
                logging.info("Continue testing videos")
                return
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(self.headers)

        # add video value
        for i, item in enumerate(self.data):
            row = [i + 1, item, '', '', '', '', '', '', '', '', '']
            self.sheet.append(row)

        self.workbook.save(self.excel_file)
        logging.info(f'Excel file "{self.excel_name}" created successfully.')

        with DB() as db:
            data_to_insert = [(item, self.version_info, self.pipline_type) for item in self.data]
            field_names = ['VIDEO', 'VERSION', 'PIPLINE_TYPE']
            db.insert_data_field('COMPATIBILITY', field_names, data_to_insert)

    # check if not video playback
    def get_videos_by_status(self, num_rows):
        last_status = self.sheet.cell(row=self.sheet.max_row, column=9).value
        if last_status == 1 or last_status == 0 or last_status == "None":
            logging.info("last row test complete")
            if len(self.left_data) > 0:
                logging.info("test left videos")
                if len(self.left_data) > num_rows:
                    videos = self.left_data[0:num_rows]
                    self.left_data = self.left_data[num_rows:]
                    return videos
                else:
                    videos = self.left_data
                    self.left_data = []
                    return videos
            else:
                return False
        else:
            max_row = self.sheet.max_row
            videos = []
            row = max_row

            while row > 2:
                status = self.sheet.cell(row=row, column=9).value
                if status == 1 or status == 0:
                    break
                row -= 1

            if row == 2:
                logging.info("first row")
            else:
                row = row + 1
            end_row = min(row + num_rows, max_row + 1)
            videos = [self.sheet.cell(row=i, column=2).value for i in range(row, end_row)]
            return videos

    # insert data to excel
    def update_excel_row(self, video_url, version_info, pipline_type, check_startPlay_flag, check_pause_flag,
                         check_resume_flag,
                         check_seek_flag, check_stopPlay_flag):
        local_pattern = r"http://[^/]+/res/test_video"
        ftp_pattern = rf"{resource_url}"
        video_filename = re.sub(local_pattern, "", video_url)
        max_row = self.sheet.max_row
        current_row = self.find_current_test_row()
        for row in range(current_row, max_row + 1):
            cell_value = self.sheet.cell(row=row, column=2).value
            cell_filename = re.sub(ftp_pattern, "", cell_value)
            if video_filename == cell_filename:
                self.sheet.cell(row=row, column=3).value = version_info
                self.sheet.cell(row=row, column=4).value = pipline_type
                self.sheet.cell(row=row, column=5).value = int(check_startPlay_flag)
                self.sheet.cell(row=row, column=6).value = int(check_pause_flag)
                self.sheet.cell(row=row, column=7).value = int(check_resume_flag)
                self.sheet.cell(row=row, column=8).value = int(check_seek_flag)
                self.sheet.cell(row=row, column=9).value = int(check_stopPlay_flag)
                status = None
                pass_value = None
                # check invalid if not
                if check_startPlay_flag == "None":
                    status = 0
                    pass_value = "invalid"
                    pass_value_font_color = 'FFFF00'
                else:
                    # check pass if not
                    if all([check_startPlay_flag, check_pause_flag, check_resume_flag, check_seek_flag,
                            check_stopPlay_flag]):
                        status = 1
                        pass_value = "pass"
                        pass_value_font_color = None
                    else:
                        status = 0
                        pass_value = "fail"
                        pass_value_font_color = 'FF0000'

                self.sheet.cell(row=row, column=10).value = pass_value
                self.sheet.cell(row=row, column=10).font = Font(color=pass_value_font_color)
                self.sheet.cell(row=row, column=11).value = status
                with DB() as db:
                    data_to_update = [int(check_startPlay_flag),
                                      int(check_pause_flag),
                                      int(check_resume_flag),
                                      int(check_seek_flag),
                                      int(check_stopPlay_flag),
                                      pass_value,
                                      status]
                    field_names = ['START', 'PAUSE', 'RESUME', 'SEEK', 'STOP', 'PASS', 'STATUS']
                    db.update_data('COMPATIBILITY', field_names, data_to_update, condition=f"VIDEO='{cell_value}'")
                break

        self.workbook.save(self.excel_file)

    def find_current_test_row(self):
        column_index = 1
        last_row_index = None
        for row in range(self.sheet.max_row, 0, -1):
            cell_value = self.sheet.cell(row=row, column=column_index).value
            if cell_value == 1:
                last_row_index = row
                logging.info(f"current ergodic row:{last_row_index}")
                break
        return last_row_index

    # close excel
    def close_workbook(self):
        self.workbook.close()
